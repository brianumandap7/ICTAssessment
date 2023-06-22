from django.shortcuts import render, HttpResponseRedirect, redirect
from django.utils import timezone
from django.contrib.auth.models import User
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin

from django.http import JsonResponse, HttpResponse

from django.core import serializers
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, F, Max, Count
from django.db import IntegrityError
from django.urls import reverse, reverse_lazy
from django.views import View
from django.views.generic import ListView, DetailView, UpdateView, CreateView
from django.views.generic.base import TemplateView
from django.db.models import Q
from datetime import date, timedelta, datetime
from django.db.models.functions import Extract, Now, Trunc
from django.db import models
from django.db.models import Sum
# from .visualizations import S_Curve_Project_Visualization, truncate

from django.core.paginator import Paginator
from django.contrib.auth.hashers import make_password
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages

from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt

from .models import Survey, office_division, CommunicationToolsOption, ProductivityOption, StorageOption, OnlineStorageOption, BackupStorageOption, OnlineOption, ICTTrainingPrograms, ProfessionalToolsOption, OfficeLocation, ConnectionOption

from .forms import demographicsform, hardwareform, SoftwareForm, CompetenciesForm, ICTTrainingsForm, StorageForm

from django.urls import reverse_lazy
from .resources import PersonResource
from tablib import Dataset
# Create your views here.
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from io import BytesIO

class view_dash(View):
    template_name = 'survey/survey.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        query = {
            'check_user': Survey.objects.filter(user=request.user),
            'sec': 1,
        }
        return render(request, self.template_name, query)

    def post(self, request, *args, **kwargs):
        if not Survey.objects.filter(user=request.user).exists():
            db = Survey()
            if request.POST.get('ia'):
                db.privacy_section = 1
                db.user = request.user
                db.save()
        
        if request.POST.get('submitted'):
            user_survey2 = Survey.objects.get(user=request.user)
            user_survey2.submitted = 1
            user_survey2.save()

        return HttpResponseRedirect('/survey/')



class simple_upload(View):
    template_name = 'survey/upload.html'

    @method_decorator(csrf_exempt)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        query = {
            'check_user': Survey.objects.filter(user=request.user),
        }
        return render(request, self.template_name, query)

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            person_resource = PersonResource()
            db = Survey()
            dataset = Dataset()

            new_person = request.FILES['myfile']

            if not new_person.name.endswith('xlsx'):
                messages.info(request, 'wrong file format')
                return render (request, '/')
            imported_data = dataset.load(new_person.read(),format = 'xlsx')
            for data in imported_data:
                username = data[1]
                if User.objects.filter(username=username).exists():
                    continue
                user = User.objects.create_user(
                    username=username,
                    email=data[3],
                    first_name=data[4],
                    last_name=data[5],
                    is_staff=data[6]
                )
                user.set_password(data[2])
                user.save()

            return HttpResponseRedirect('/survey/')
        

        db.save()

        return HttpResponseRedirect('/survey/')

class demographics(UpdateView):
    model = Survey
    form_class = demographicsform
    template_name = 'survey/demographics.html'
    success_url = reverse_lazy('survey:dash')

    def get_context_data(self, **kwargs):
        context = super(demographics, self).get_context_data(**kwargs)
        context['title'] = 'Demographics'
        context['check_user'] = Survey.objects.filter(user = self.request.user)
        context['sec'] = 2
        context['pto'] = OfficeLocation.objects.count() - 2
        context['olo'] = OfficeLocation.objects.count() - 1
        context['pk'] = self.kwargs.get('pk')

        return context

class profile(View):
    template_name = 'survey/profile.html'


    def get(self, request, *args, **kwargs):
        query = {
            'check_user': Survey.objects.filter(user=request.user),
        }

        return render(request, self.template_name, query)

class hardware(UpdateView):
    model = Survey
    form_class = hardwareform
    template_name = 'survey/hardware.html'
    success_url = reverse_lazy('survey:dash')

    def get_context_data(self, **kwargs):
        context = super(hardware, self).get_context_data(**kwargs)
        context['title'] = 'Hardware'
        context['check_user'] = Survey.objects.filter(user = self.request.user)
        context['sec'] = 3
        context['pk'] = self.kwargs.get('pk')

        return context

class Software(LoginRequiredMixin, UpdateView):
    model = Survey
    form_class = SoftwareForm
    template_name = 'survey/software.html'
    success_url = reverse_lazy('survey:dash')
    
    def get_context_data(self, **kwargs):
        context = super(Software, self).get_context_data(**kwargs)
        context['title'] = 'Software'
        context['check_user'] = Survey.objects.filter(user = self.request.user)
        context['sec'] = 4
        context['pk'] = self.kwargs.get('pk')
        context['pto'] = ProfessionalToolsOption.objects.count() - 1
        context['ct'] = CommunicationToolsOption.objects.count() - 1
        context['ps'] = ProductivityOption.objects.count() - 1
        return context

class Storage(LoginRequiredMixin, UpdateView):
    model = Survey
    form_class = StorageForm
    template_name = 'survey/storage.html'
    success_url = reverse_lazy('survey:dash')
    
    def get_context_data(self, **kwargs):
        context = super(Storage, self).get_context_data(**kwargs)
        context['title'] = 'Storage'
        context['check_user'] = Survey.objects.filter(user = self.request.user)
        context['sec'] = 5
        context['pk'] = self.kwargs.get('pk')
        context['scount'] = StorageOption.objects.count() - 1
        context['bcount'] = BackupStorageOption.objects.count() - 1
        return context

class Competencies(LoginRequiredMixin, UpdateView):
    model = Survey
    form_class = CompetenciesForm
    template_name = 'survey/competencies.html'
    success_url = reverse_lazy('survey:dash')

    def get_context_data(self, **kwargs):
        context = super(Competencies, self).get_context_data(**kwargs)
        context['title'] = 'Competencies'
        context['check_user'] = Survey.objects.filter(user = self.request.user)
        context['sec'] = 6
        context['pk'] = self.kwargs.get('pk')
        return context


class ICTTrainings(LoginRequiredMixin, UpdateView):
    model = Survey
    form_class = ICTTrainingsForm
    template_name = 'survey/ict_trainings.html'
    success_url = reverse_lazy('survey:dash')

    def get_context_data(self, **kwargs):
        context = super(ICTTrainings, self).get_context_data(**kwargs)
        context['title'] = 'ICT Trainings'
        context['check_user'] = Survey.objects.filter(user = self.request.user)
        context['sec'] = 7
        context['pk'] = self.kwargs.get('pk')
        context['ict_training_count'] = ICTTrainingPrograms.objects.count() - 1
        context['one'] = 'Training on fundamental abilities and knowledge that are essential for effectively using a computer and its associated software.'
        context['two'] = 'The training focuses on equipping individuals with the competencies to engage with digital tools, platforms, and resources in various contexts.'
        context['three'] = 'Cybersecurity awareness training promotes online safety and empowers individuals to protect themselves and their organizations from cyber threats.'
        context['four'] = 'Data analysis training equips individuals with the skills to effectively analyze data, derive insights, and make informed decisions.'
        context['five'] = 'Digital marketing training imparts the skills needed to succeed in promoting products and brands through online channels.'
        context['six'] = 'Cloud computing training imparts the necessary skills and knowledge to effectively utilize and manage cloud-based technologies and services.'
        context['seven'] = 'IoT training imparts skills for developing and managing connected devices, harnessing the potential of IoT technologies.'
        context['eight'] = 'Website design and development training builds skills for creating functional and visually appealing websites.'
        context['nine'] = 'Digital Productivity Tools training boosts efficiency by equipping individuals with skills to leverage various digital tools effectively.'
        context['ten'] = 'Software Testing training ensures software quality through effective testing techniques.'
        context['eleven'] = 'SQA training ensures high-quality software by providing individuals with the necessary skills and knowledge for effective quality assurance practices.' 
        context['twelve'] = 'SAD training equips individuals with the skills to analyze and design efficient systems for solving complex business problems.'
        context['thirteen'] = 'I.T. Project Management training equips individuals with the skills to effectively plan, execute, and oversee projects in the field of information technology.'
        context['fourteen'] = 'Digital Innovation training enhances skills for leveraging emerging technologies and creative strategies in the digital era.'
        context['fifteen'] = 'Digital Transformation training develops skills for effectively navigating and driving organizational change in the digital era.'
        context['sixteen'] = 'ICT Policy Formulation and Development training enables effective ICT policy creation and implementation.'
        return context

def export_to_excel(request):
    surveys = Survey.objects.filter(submitted = 1)

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Add column headers
    headers = ['ID', 'Username', 'Office', 'Presidential Appointee', 
    'Permanent', 'Co-terminous', 'JO and COS', 'Casual Temporary',
    'Male', 'Female', 'age 20 to 24', 'age 25 to 34', 'age 35 to 44',
    'age 45 to 54', 'age 55 and above',
    'Office Locations', 'PMO Office', 'Other Office',
    'Desktop Acer Installed', 'Desktop HP Installed','Desktop Lenovo Installed', 'Desktop Samsung Installed', 'Desktop Others Installed',
    'Desktop Acer personal', 'desktop hp personal', 'desktop lenovo personal', 'dekstop samsung personal', 'desktop others personal',
    'tablet_acer_installed', 'tablet_hp_installed', 'tablet_lenovo_installed', 'tablet_samsung_installed', 'tablet_others_installed',
    'tablet_acer_personal', 'tablet_hp_personal', 'tablet_lenovo_personal', 'tablet_samsung_personal', 'tablet_others_personal',
    'mouse_acer_installed', 'mouse_hp_installed', 'mouse_lenovo_installed', 'mouse_samsung_installed', 'mouse_others_installed',
    'mouse_acer_personal', 'mouse_hp_personal', 'mouse_lenovo_personal', 'mouse_samsung_personal', 'mouse_others_personal',
    'kb_acer_installed', 'kb_hp_installed', 'kb_lenovo_installed', 'kb_samsung_installed', 'kb_others_installed',
    'kb_acer_personal', 'kb_hp_personal', 'kb_lenovo_personal', 'kb_samsung_personal', 'kb_others_personal',
    'monitor_acer_installed', 'monitor_hp_installed', 'monitor_lenovo_installed', 'monitor_samsung_installed', 'monitor_others_installed',
    'monitor_acer_personal', 'monitor_hp_personal', 'monitor_lenovo_personal', 'monitor_samsung_personal', 'monitor_others_personal',
    'laptop_acer_installed', 'laptop_hp_installed', 'laptop_lenovo_installed', 'laptop_samsung_installed', 'laptop_others_installed',
    'laptop_acer_personal', 'laptop_hp_personal', 'laptop_lenovo_personal', 'laptop_samsung_personal', 'laptop_others_personal',
    'Internet Connection Setup',
    'Ordinaryprinter shared', 'Coloredprinter shared', 'Scanner shared', 'Speaker shared', 'Tv shared', 'Projector shared',
    'Camera shared', 'Microphone shared', 'Photocopier shared',
    'Ordinaryprinter shared Personal', 'Coloredprinter shared Personal', 'Scanner shared Personal', 'Speaker shared Personal', 'Tv shared Personal', 'Projector shared Personal',
    'Camera shared Personal', 'Microphone shared Personal', 'Photocopier shared Personal',
    'desktop need', 'laptop need', 'tablet need', 'mouse need', 'kb need', 'monitor need', 'ordinaryprinter need', 'coloredprinter need',
    'scanner need', 'speaker need', 'tv need', 'projector need', 'camera need', 'microphone need', 'photocopier need', 'hardware comments',
    'dotr_issued_professional_tools_installed','dotr_issued_professional_tools_installed_others',
    'communication_tools_installed', 'communication_tools_installed_others', 'communication_tools_needed', 'communication_tools_needed_others',
    'communication_frequent', 'communication_most_needed','Personal owned professional tools installed', 'Personal owned professional tools installed others',
    'professional tools needed', 'professional tools needed others', 'Professional Frequent', 'Professional Most Needed',
    'DOTr productivity', 'DOTr productivity others', 'Personal productivity', 'Personal productivity others', 'Productivity Needed', 'Productivity Needed Others',
    'Productivity Frequent', 'Productivity Most Needed', 'Software Comments', 'DOTR Storage', 'DOTR Storage others', 'Personal Storage', 'Personal Storage Others',
    'Storage Need', 'Storage Need Others', 'DOTR Online Storage', 'Personal Online Storage', 'Online Storage Need', 'Backup Storage', 'Backup Storage Others',
    'Backup Storage Need', 'Backup Storage Need Others', 'Storage Comments',
    'ICT Trainings Taken', 'ICT Trainings Taken Others', 'ICT Trainins Interests', 'ICT Trainings Interests Others', 'ICT Training Comments',
    'Turning On Familiar',
    'Network Cable Familiar',
    'Network WiFi Familiar',
    'USB Printer Familiar',
    'Wireless Printer Familiar',
    'Saving Files Familiar',
    'Creating Folders Familiar',
    'Copying Deleting Familiar',
    'Installing Software',
    'File Types Familiar',
    'Unzip Familiar',
    'File Search Familiar',
    'Connect Device Familiar',
    'Access Email Familiar',
    'Sending Email Familiar',
    'Add Attachments Familiar',
    'Add Signatures Familiar',
    'Mailing List Familiar',
    'Gmail Familiar',
    'Meet Familiar',
    'Calendar Familiar',
    'Drive Familiar',
    'Docs Familiar',
    'Sheets Familiar',
    'Sliders Familiar',
    'Forms Familiar',
    'Sites Familiar',
    'Keep Familiar',
    'Outlook Familiar',
    'Teams Familiar',
    'OneDrive Familiar',
    'Word Familiar',
    'Excel Familiar',
    'PPT Familiar',
    'Forms2 Familiar',
    'SharePoint Familiar',
    'OneNote Familiar',
    'Power BI Familiar',
    'Online Options',
    'Competencies Comments',
    'Submitted',
    ]

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = Font(bold=True)

    # Add survey data
    for row_num, survey in enumerate(surveys, 2):
        ws.cell(row=row_num, column=1, value=survey.survey_id)
        user = survey.user.username if survey.user else ''
        username = survey.user.username if survey.user else ''
        ws.cell(row=row_num, column=2, value=username)
        ws.cell(row=row_num, column=3, value=str(survey.office_division_name))
        ws.cell(row=row_num, column=4, value=survey.presidential_appointee)
        ws.cell(row=row_num, column=5, value=survey.permanent)
        ws.cell(row=row_num, column=6, value=survey.coterminus)
        ws.cell(row=row_num, column=7, value=survey.jo_cos)
        ws.cell(row=row_num, column=8, value=survey.casual_temporary)
        ws.cell(row=row_num, column=9, value=survey.male)
        ws.cell(row=row_num, column=10, value=survey.female)
        ws.cell(row=row_num, column=11, value=survey.age_20_24)
        ws.cell(row=row_num, column=12, value=survey.age_25_34)
        ws.cell(row=row_num, column=13, value=survey.age_35_44)
        ws.cell(row=row_num, column=14, value=survey.age_45_54)
        ws.cell(row=row_num, column=15, value=survey.age_55_above)
        office_location = ', '.join([location.title for location in survey.office_location.all()])
        ws.cell(row=row_num, column=16, value=office_location)
        ws.cell(row=row_num, column=17, value=survey.pmo_office)
        ws.cell(row=row_num, column=18, value=survey.office_location_others)

        ws.cell(row=row_num, column=19, value=survey.desktop_acer_installed)
        ws.cell(row=row_num, column=20, value=survey.desktop_hp_installed)
        ws.cell(row=row_num, column=21, value=survey.desktop_lenovo_installed)
        ws.cell(row=row_num, column=22, value=survey.desktop_samsung_installed)
        ws.cell(row=row_num, column=23, value=survey.desktop_others_installed)

        ws.cell(row=row_num, column=24, value=survey.desktop_acer_personal)
        ws.cell(row=row_num, column=25, value=survey.desktop_hp_personal)
        ws.cell(row=row_num, column=26, value=survey.desktop_lenovo_personal)
        ws.cell(row=row_num, column=27, value=survey.desktop_samsung_personal)
        ws.cell(row=row_num, column=28, value=survey.desktop_others_personal)

        ws.cell(row=row_num, column=29, value=survey.tablet_acer_installed)
        ws.cell(row=row_num, column=30, value=survey.tablet_hp_installed)
        ws.cell(row=row_num, column=31, value=survey.tablet_lenovo_installed)
        ws.cell(row=row_num, column=32, value=survey.tablet_samsung_installed)
        ws.cell(row=row_num, column=33, value=survey.tablet_others_installed)

        ws.cell(row=row_num, column=34, value=survey.tablet_acer_personal)
        ws.cell(row=row_num, column=35, value=survey.tablet_hp_personal)
        ws.cell(row=row_num, column=36, value=survey.tablet_lenovo_personal)
        ws.cell(row=row_num, column=37, value=survey.tablet_samsung_personal)
        ws.cell(row=row_num, column=38, value=survey.tablet_others_personal)

        ws.cell(row=row_num, column=39, value=survey.mouse_acer_installed)
        ws.cell(row=row_num, column=40, value=survey.mouse_hp_installed)
        ws.cell(row=row_num, column=41, value=survey.mouse_lenovo_installed)
        ws.cell(row=row_num, column=42, value=survey.mouse_samsung_installed)
        ws.cell(row=row_num, column=43, value=survey.mouse_others_installed)

        ws.cell(row=row_num, column=44, value=survey.mouse_acer_personal)
        ws.cell(row=row_num, column=45, value=survey.mouse_hp_personal)
        ws.cell(row=row_num, column=46, value=survey.mouse_lenovo_personal)
        ws.cell(row=row_num, column=47, value=survey.mouse_samsung_personal)
        ws.cell(row=row_num, column=48, value=survey.mouse_others_personal)

        ws.cell(row=row_num, column=49, value=survey.kb_acer_installed)
        ws.cell(row=row_num, column=50, value=survey.kb_hp_installed)
        ws.cell(row=row_num, column=51, value=survey.kb_lenovo_installed)
        ws.cell(row=row_num, column=52, value=survey.kb_samsung_installed)
        ws.cell(row=row_num, column=53, value=survey.kb_others_installed)

        ws.cell(row=row_num, column=54, value=survey.kb_acer_personal)
        ws.cell(row=row_num, column=55, value=survey.kb_hp_personal)
        ws.cell(row=row_num, column=56, value=survey.kb_lenovo_personal)
        ws.cell(row=row_num, column=57, value=survey.kb_samsung_personal)
        ws.cell(row=row_num, column=58, value=survey.kb_others_personal)

        ws.cell(row=row_num, column=59, value=survey.monitor_acer_installed)
        ws.cell(row=row_num, column=60, value=survey.monitor_hp_installed)
        ws.cell(row=row_num, column=61, value=survey.monitor_lenovo_installed)
        ws.cell(row=row_num, column=62, value=survey.monitor_samsung_installed)
        ws.cell(row=row_num, column=63, value=survey.monitor_others_installed)

        ws.cell(row=row_num, column=64, value=survey.monitor_acer_personal)
        ws.cell(row=row_num, column=65, value=survey.monitor_hp_personal)
        ws.cell(row=row_num, column=66, value=survey.monitor_lenovo_personal)
        ws.cell(row=row_num, column=67, value=survey.monitor_samsung_personal)
        ws.cell(row=row_num, column=68, value=survey.monitor_others_personal)

        ws.cell(row=row_num, column=69, value=survey.laptop_acer_installed)
        ws.cell(row=row_num, column=70, value=survey.laptop_hp_installed)
        ws.cell(row=row_num, column=71, value=survey.laptop_lenovo_installed)
        ws.cell(row=row_num, column=72, value=survey.laptop_samsung_installed)
        ws.cell(row=row_num, column=73, value=survey.laptop_others_installed)

        ws.cell(row=row_num, column=74, value=survey.laptop_acer_personal)
        ws.cell(row=row_num, column=75, value=survey.laptop_hp_personal)
        ws.cell(row=row_num, column=76, value=survey.laptop_lenovo_personal)
        ws.cell(row=row_num, column=77, value=survey.laptop_samsung_personal)
        ws.cell(row=row_num, column=78, value=survey.laptop_others_personal)

        connection_option = ', '.join([connection.title for connection in survey.connection_option.all()])
        ws.cell(row=row_num, column=79, value=connection_option)


        ws.cell(row=row_num, column=80, value=survey.ordinaryprinter_shared)
        ws.cell(row=row_num, column=81, value=survey.coloredprinter_shared)
        ws.cell(row=row_num, column=82, value=survey.scanner_shared)
        ws.cell(row=row_num, column=83, value=survey.speaker_shared)
        ws.cell(row=row_num, column=84, value=survey.tv_shared)
        ws.cell(row=row_num, column=85, value=survey.projector_shared)
        ws.cell(row=row_num, column=86, value=survey.camera_shared)
        ws.cell(row=row_num, column=87, value=survey.microphone_shared)
        ws.cell(row=row_num, column=88, value=survey.photocopier_shared)

        ws.cell(row=row_num, column=89, value=survey.ordinaryprinter_shared_personal)
        ws.cell(row=row_num, column=90, value=survey.coloredprinter_shared_personal)
        ws.cell(row=row_num, column=91, value=survey.scanner_shared_personal)
        ws.cell(row=row_num, column=92, value=survey.speaker_shared_personal)
        ws.cell(row=row_num, column=93, value=survey.tv_shared_personal)
        ws.cell(row=row_num, column=94, value=survey.projector_shared_personal)
        ws.cell(row=row_num, column=95, value=survey.camera_shared_personal)
        ws.cell(row=row_num, column=96, value=survey.microphone_shared_personal)
        ws.cell(row=row_num, column=97, value=survey.photocopier_shared_personal)

        ws.cell(row=row_num, column=98, value=survey.desktop_need)
        ws.cell(row=row_num, column=99, value=survey.laptop_need)
        ws.cell(row=row_num, column=100, value=survey.tablet_need)
        ws.cell(row=row_num, column=101, value=survey.mouse_need)
        ws.cell(row=row_num, column=102, value=survey.keyboard_need)
        ws.cell(row=row_num, column=103, value=survey.monitor_need)
        ws.cell(row=row_num, column=104, value=survey.ordinaryprinter_need)
        ws.cell(row=row_num, column=105, value=survey.coloredprinter_need)
        ws.cell(row=row_num, column=106, value=survey.scanner_need)
        ws.cell(row=row_num, column=107, value=survey.speaker_need)
        ws.cell(row=row_num, column=108, value=survey.tv_need)
        ws.cell(row=row_num, column=109, value=survey.projector_need)
        ws.cell(row=row_num, column=110, value=survey.camera_need)
        ws.cell(row=row_num, column=111, value=survey.microphone_need)
        ws.cell(row=row_num, column=112, value=survey.photocopier_need)

        ws.cell(row=row_num, column=113, value=survey.hardware_comments)

        dotr_issued_professional_tools_installed = ', '.join([professional.title for professional in survey.dotr_issued_professional_tools_installed.all()])
        ws.cell(row=row_num, column=114, value=dotr_issued_professional_tools_installed)

        ws.cell(row=row_num, column=115, value=str(survey.dotr_issued_professional_tools_installed_others))

        communication_tools_installed = ', '.join([communication.title for communication in survey.communication_tools_installed.all()])
        ws.cell(row=row_num, column=116, value=communication_tools_installed)

        ws.cell(row=row_num, column=117, value=str(survey.communication_tools_installed_others))

        communication_tools_needed = ', '.join([communication_needed.title for communication_needed in survey.communication_tools_needed.all()])
        ws.cell(row=row_num, column=118, value=communication_tools_needed)

        ws.cell(row=row_num, column=119, value=str(survey.communication_tools_needed_others))
        ws.cell(row=row_num, column=120, value=str(survey.communication_frequent))
        ws.cell(row=row_num, column=121, value=str(survey.communication_most_needed))

        personal_owned_professional_tools_installed = ', '.join([personal_ins.title for personal_ins in survey.personal_owned_professional_tools_installed.all()])
        ws.cell(row=row_num, column=122, value=personal_owned_professional_tools_installed)

        ws.cell(row=row_num, column=123, value=str(survey.personal_owned_professional_tools_installed_others))

        professional_tools_needed = ', '.join([prof.title for prof in survey.professional_tools_needed.all()])
        ws.cell(row=row_num, column=124, value=professional_tools_needed)

        ws.cell(row=row_num, column=125, value=str(survey.professional_tools_needed_others))
        ws.cell(row=row_num, column=126, value=str(survey.professional_frequent))
        ws.cell(row=row_num, column=127, value=str(survey.professional_most_needed))

        dotr_issued_productivity_installed = ', '.join([prodi.title for prodi in survey.dotr_issued_productivity_installed.all()])
        ws.cell(row=row_num, column=128, value=dotr_issued_productivity_installed)
        ws.cell(row=row_num, column=129, value=str(survey.dotr_issued_productivity_installed_others))

        personal_owned_productivity_installed = ', '.join([prodp.title for prodp in survey.personal_owned_productivity_installed.all()])
        ws.cell(row=row_num, column=130, value=personal_owned_productivity_installed)
        ws.cell(row=row_num, column=131, value=str(survey.personal_owned_productivity_installed_others))

        productivity_needed = ', '.join([pn.title for pn in survey.productivity_needed.all()])
        ws.cell(row=row_num, column=132, value=productivity_needed)
        ws.cell(row=row_num, column=133, value=str(survey.productivity_needed_others))

        ws.cell(row=row_num, column=134, value=str(survey.productivity_frequent))
        ws.cell(row=row_num, column=135, value=str(survey.productivity_most_needed))

        ws.cell(row=row_num, column=136, value=str(survey.software_comments))

        dotr_storage = ', '.join([ds.title for ds in survey.dotr_storage.all()])
        ws.cell(row=row_num, column=137, value=dotr_storage)
        ws.cell(row=row_num, column=138, value=str(survey.dotr_sotrage_others))

        personal_storage = ', '.join([ps.title for ps in survey.personal_storage.all()])
        ws.cell(row=row_num, column=139, value=personal_storage)
        ws.cell(row=row_num, column=140, value=str(survey.personal_storage_others))

        storage_need = ', '.join([sn.title for sn in survey.storage_need.all()])
        ws.cell(row=row_num, column=141, value=storage_need)
        ws.cell(row=row_num, column=142, value=str(survey.storage_need_others))

        dotr_online_storage = ', '.join([sn.title for sn in survey.dotr_online_storage.all()])
        ws.cell(row=row_num, column=143, value=dotr_online_storage)

        personal_online_storage = ', '.join([sn.title for sn in survey.personal_online_storage.all()])
        ws.cell(row=row_num, column=144, value=personal_online_storage)

        online_storage_need = ', '.join([sn.title for sn in survey.online_storage_need.all()])
        ws.cell(row=row_num, column=145, value=online_storage_need)

        backup_storage = ', '.join([sn.title for sn in survey.backup_storage.all()])
        ws.cell(row=row_num, column=146, value=backup_storage)
        ws.cell(row=row_num, column=147, value=str(survey.backup_storage_others))

        backup_storage_need = ', '.join([sn.title for sn in survey.backup_storage_need.all()])
        ws.cell(row=row_num, column=148, value=backup_storage_need)
        ws.cell(row=row_num, column=149, value=str(survey.backup_storage_need_others))

        ws.cell(row=row_num, column=150, value=str(survey.storage_comments))

        ict_trainings_taken = ', '.join([sn.title for sn in survey.ict_trainings_taken.all()])
        ws.cell(row=row_num, column=151, value=ict_trainings_taken)
        ws.cell(row=row_num, column=152, value=str(survey.ict_trainings_taken_others))

        ict_trainings_interests = ', '.join([sn.title for sn in survey.ict_trainings_interests.all()])
        ws.cell(row=row_num, column=153, value=ict_trainings_interests)
        ws.cell(row=row_num, column=154, value=str(survey.ict_trainings_interests_others))
        
        ws.cell(row=row_num, column=155, value=str(survey.ict_training_comments))

        ws.cell(row=row_num, column=156, value=str(survey.turning_on_familiar))
        ws.cell(row=row_num, column=157, value=str(survey.network_cable_familiar))
        ws.cell(row=row_num, column=158, value=str(survey.network_wifi_familiar))
        ws.cell(row=row_num, column=159, value=str(survey.usb_printer_familiar))
        ws.cell(row=row_num, column=160, value=str(survey.wireless_printer_familiar))
        ws.cell(row=row_num, column=161, value=str(survey.saving_files_familiar))
        ws.cell(row=row_num, column=162, value=str(survey.creating_folders_familiar))
        ws.cell(row=row_num, column=163, value=str(survey.copying_deleting_familiar))
        ws.cell(row=row_num, column=164, value=str(survey.installing_software))
        ws.cell(row=row_num, column=165, value=str(survey.file_types_familiar))
        ws.cell(row=row_num, column=166, value=str(survey.unzip_familiar))
        ws.cell(row=row_num, column=167, value=str(survey.file_search_familiar))
        ws.cell(row=row_num, column=168, value=str(survey.connect_device_familiar))
        ws.cell(row=row_num, column=169, value=str(survey.access_email_familiar))
        ws.cell(row=row_num, column=170, value=str(survey.sending_email_familiar))
        ws.cell(row=row_num, column=171, value=str(survey.add_attachments_familiar))
        ws.cell(row=row_num, column=172, value=str(survey.add_signatures_familiar))
        ws.cell(row=row_num, column=173, value=str(survey.mailing_list_familiar))
        ws.cell(row=row_num, column=174, value=str(survey.gmail_familiar))
        ws.cell(row=row_num, column=175, value=str(survey.meet_familiar))
        ws.cell(row=row_num, column=176, value=str(survey.calendar_familiar))
        ws.cell(row=row_num, column=177, value=str(survey.drive_familiar))
        ws.cell(row=row_num, column=178, value=str(survey.docs_familiar))
        ws.cell(row=row_num, column=179, value=str(survey.sheets_familiar))
        ws.cell(row=row_num, column=180, value=str(survey.sliders_familiar))
        ws.cell(row=row_num, column=181, value=str(survey.forms_familiar))
        ws.cell(row=row_num, column=182, value=str(survey.sites_familiar))
        ws.cell(row=row_num, column=183, value=str(survey.keep_familiar))
        ws.cell(row=row_num, column=184, value=str(survey.outlook_familiar))
        ws.cell(row=row_num, column=185, value=str(survey.teams_familiar))
        ws.cell(row=row_num, column=186, value=str(survey.onedrive_familiar))
        ws.cell(row=row_num, column=187, value=str(survey.word_familiar))
        ws.cell(row=row_num, column=188, value=str(survey.excel_familiar))
        ws.cell(row=row_num, column=189, value=str(survey.ppt_familiar))
        ws.cell(row=row_num, column=190, value=str(survey.forms2_familiar))
        ws.cell(row=row_num, column=191, value=str(survey.sharepoint_familiar))
        ws.cell(row=row_num, column=192, value=str(survey.onenote_familiar))
        ws.cell(row=row_num, column=193, value=str(survey.powerbi_familiar))
        ws.cell(row=row_num, column=194, value=str(survey.online_options))
        ws.cell(row=row_num, column=195, value=str(survey.competencies_comments))

        ws.cell(row=row_num, column=196, value=str(survey.submitted))

    # Set column widths
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15

    # Save the workbook to a BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Create the HttpResponse object with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=surveys.xlsx'
    response.write(buffer.getvalue())

    return response
